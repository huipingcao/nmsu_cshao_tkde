import os
from os.path import isfile, join, isdir
import numpy as np
from datetime import datetime
import xlwt
import xlrd
import openpyxl

################################################################
#File IO

def check_file_exists(data_file):
    return isfile(data_file)

def listall_withinsubdir(dir):
    return os.walk(dir)

def list_files(dir):
    onlyfiles = [f for f in os.listdir(dir) if isfile(join(dir, f))]
    return sorted(onlyfiles)

def listall(dir):
    return os.listdir(dir)

def list_sub_folders(dir):
    sub_folders = [f for f in os.listdir(dir) if isdir(join(dir,f))]
    return sorted(sub_folders)


def delete_files_startwith(dir, start_str):
    all_files = list_files(dir)
    for file in all_files:
        if file.startswith(start_str):
            os.remove(os.path.join(dir, file))

def check_folder_exists(data_folder):
    try:
        os.stat(data_folder)
    except:
        print data_folder + ' does not exists!'
        return False
    return True

def check_file_exists(data_file):
    return isfile(data_file)


# Check folder exists or not, create folders if not exists
def init_folder(data_folder):
    split_key = '/'
    if data_folder.endswith(split_key):
        data_folder = data_folder[:-1]
    folder_array = data_folder.split(split_key)
    data_folder = ''
    for item in folder_array:
        data_folder = data_folder + item + split_key
        if item == '..':
            continue
        try:
            os.stat(data_folder)
        except:
            os.mkdir(data_folder)

    return data_folder




#End of file IO
################################################################

def file_read_split(file_name, class_column=0, delimiter=' ', header=True):
    data_matrix, attr_num = file_reading(file_name, delimiter, header)
    x_matrix, y_vector = x_y_spliting(data_matrix, class_column)
    x_row, x_col = x_matrix.shape
    attr_len = x_col/attr_num
    x_matrix = x_matrix.reshape(x_row, attr_num, attr_len)
    return x_matrix, y_vector


##
# Read the giving file and store the data into matrix structure
# the return format is folat, and the minimal value of y may not be 0
def file_reading(file_name, delimiter=' ', header=True):
    num = 0
    data_matrix = []
    header_line = "-1"
    with open(file_name) as f:
        data_row = []
        for line in f:
            if header == True:
                header = False
                header_line = line.strip()
                continue
            num = num + 1
            #if num > 100:
            #    break
            data_row = line.split(delimiter)
            data_matrix.append(data_row)
    row_num = len(data_matrix)
    col_num = len(data_matrix[0])
    data_matrix = np.array(data_matrix, dtype=float)#.reshape(row_num, col_num)
    data_matrix.astype(float)
    header_line = int(header_line)
    return data_matrix, header_line

##
# Read the giving file and store the data into matrix structure
# the return format is folat, and the minimal value of y may not be 0
def file_writing(data_matrix, file_name, attr_num=-1, delimiter=' '):
    data_row, data_col = data_matrix.shape
    with open(file_name, 'w') as f:
        if attr_num > 0:
            f.write(str(int(attr_num)) + '\n')
        for row in range(0, data_row):
            row_vector = data_matrix[row, :]
            row_str = str(int(row_vector[0]))
            for index in range(1, data_col):
                row_str = row_str + delimiter + str(row_vector[index])
            f.write(row_str + '\n')
    

##
# from a data_matrix, split the x_matrix and y_vector based on class_column
# values in y_vector can not be negative
# returned y_vector with minimal value is 0
def x_y_spliting(data_matrix, class_column):
    y_vector = data_matrix[:, class_column].astype(int)
    y_vector = y_vector - min(y_vector)
    x_matrix = np.delete(data_matrix, class_column, 1)
    return x_matrix, y_vector



def train_test_file_reading(train_file, test_file, class_column=0, delimiter=' ', header=True):
    train_matrix, attr_num = file_reading(train_file, delimiter, header)
    test_matrix, attr_num = file_reading(test_file, delimiter, header)
    train_x_matrix, train_y_vector = x_y_spliting(train_matrix, class_column)
    test_x_matrix, test_y_vector = x_y_spliting(test_matrix, class_column)
    train_min_class = min(train_y_vector)
    test_min_class = min(test_y_vector)
    if train_min_class != 0 or test_min_class !=0:
        raise Exception("minimum class does not match")
    return train_x_matrix, train_y_vector, test_x_matrix, test_y_vector

def train_test_file_reading_with_attrnum(train_file, test_file, class_column=0, delimiter=' ', header=True):
    train_matrix, attr_num = file_reading(train_file, delimiter, header)
    test_matrix, attr_num = file_reading(test_file, delimiter, header)
    train_x_matrix, train_y_vector = x_y_spliting(train_matrix, class_column)
    test_x_matrix, test_y_vector = x_y_spliting(test_matrix, class_column)
    train_min_class = min(train_y_vector)
    test_min_class = min(test_y_vector)
    if train_min_class != 0 or test_min_class !=0:
        raise Exception("minimum class does not match")
    return train_x_matrix, train_y_vector, test_x_matrix, test_y_vector, int(attr_num)




def write_to_excel(data_matrix, excel_file, start_row=1, start_col=1, sheet_name='sheet1'):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.get_sheet_by_name(sheet_name)
    except:
        print "!!!"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        #worksheet = workbook.create_sheet(sheet_name)
    
    for row, row_value in enumerate(data_matrix):
        for col, col_value in enumerate(row_value):
            try:
                col_value = float(col_value)
            except:
                pass
            worksheet.cell(row=row+start_row, column=col+start_col).value = col_value
    workbook.save(excel_file)



def data_dict_loading(data_folder, data_keyword, class_column=0, delimiter=' ', header=True):
    data_file_list = list_files(data_folder)
    data_dict = {}
    for data_file in data_file_list:
        if data_keyword not in data_file:
            continue
        fold_num = int(data_file.split('_')[1])
        train_matrix, attr_num = file_reading(data_folder + data_file, delimiter, header)
        train_x_matrix, train_y_vector = x_y_spliting(train_matrix, class_column)
        data_dict[fold_num] = [train_x_matrix, train_y_vector]
    return data_dict


if __name__ == '__main__':
    #data_folder = '../../data/arc_activity_recognition/s1_ijcal_10folds/'
    #file_name_1 = "train.txt"
    #file_name_2 = "test.txt"
    #out_file = "all.txt"
    #merge_two_files(data_folder, file_name_1, file_name_2, out_file)

    obj_folder = "../../object/dsa/train_test_3_fold/"
    print list_sub_folders(obj_folder)
    sdf
    attr_num = 2
    attr_len = 3
    a = np.random.rand(2, attr_num*attr_len)
    b = train_test_transpose(a, attr_num, attr_len)
    print "a"
    print a.reshape(2, attr_num, attr_len)
    print "b"
    print b.reshape(2, attr_len, attr_num)

